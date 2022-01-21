import os
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def createTargetIfNotExists(tgt_workbook, tgt_worksheet):

  targetW = Workbook()
  try:
    targetW = load_workbook(tgt_workbook)
  except:
    targetW.save(tgt_workbook)
    targetW = load_workbook(tgt_workbook)

  sheets = targetW.sheetnames

  if tgt_worksheet not in sheets:
    print("Creating sheet: ", tgt_worksheet)
    targetW.create_sheet(tgt_worksheet)
    targetW.save(tgt_workbook)

def getRangeOfRowCols(xlRange):

  rangeTuples = []

  start,end = xlRange.split(":")
  startCoord = coordinate_from_string(start)
  startCol, startRow = column_index_from_string(startCoord[0]), startCoord[1]
  endCoord = coordinate_from_string(end)
  endCol, endRow =  column_index_from_string(endCoord[0]), endCoord[1]

  rangeTuples.append((startRow, endRow))
  rangeTuples.append((startCol, endCol))

  return rangeTuples

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def buildTargetRange(srcCoords, tr):

  rangeTuples = []

  startCoord = coordinate_from_string(tr)
  startCol, startRow = column_index_from_string(startCoord[0]), startCoord[1]
  endCol, endRow = startCol + srcCoords[1][1] - srcCoords[1][0], startRow + srcCoords[0][1] - srcCoords[0][0]
  rangeTuples.append((startRow, endRow))
  rangeTuples.append((startCol, endCol))

  return rangeTuples     

def getNewTargetCell(tr, rows_to_be_added):

  x = coordinate_from_string(tr)
  y = x[1] + rows_to_be_added
  return x[0] + str(y)

def copyCellsFromSrcToTgt(sw, tw, tws, tr, ml):

  sWb = load_workbook(sw)
  tWb = load_workbook(tw)

  for m in ml:
    srcSheet = sWb[m[0]]
    tgtSheet = tWb[tws]

    srcRangeDetails = getRangeOfRowCols(m[1])
    tgtRangeDetails = buildTargetRange(srcRangeDetails, tr)

    rows_to_be_added = srcRangeDetails[0][1] - srcRangeDetails[0][0] + 1

    rangeSelected = copyRange(srcRangeDetails[1][0], srcRangeDetails[0][0], srcRangeDetails[1][1], srcRangeDetails[0][1], srcSheet)
    if m[2]:
      tgtSheet.insert_rows(tgtRangeDetails[0][0]+1, amount = rows_to_be_added)
    pasteRange(tgtRangeDetails[1][0], tgtRangeDetails[0][0], tgtRangeDetails[1][1], tgtRangeDetails[0][1], tgtSheet, rangeSelected)

    tr = getNewTargetCell(tr, rows_to_be_added)
    
  tWb.save(tw)

def returnMapFileDetails(mapping_file):

  map_details = []

  wb = load_workbook(mapping_file)
  ws = wb["Sheet1"]

  for r1 in range(1, ws.max_row + 1):
      if ws.cell(row=r1, column=1).value is not None:
        r = (ws.cell(row=r1, column=1).value, ws.cell(row=r1, column=2).value, ws.cell(row=r1, column=3).value)
        map_details.append(r)

  return map_details

if __name__ == "__main__":
  #src_workbook = input("Enter path and name of src workbook: ")
  #src_worksheet = input("Enter name of the src sheet: ")
  #src_range = input("Enter the range of src cells: ")

  #tgt_workbook = input("Enter path and name of target workbook: ")
  #tgt_worksheet = input("Enter name of the target sheet: ")
  #tgt_range = input("Enter the range of target cells: ")

  # create parser
  parser = argparse.ArgumentParser()

  # add arguments to the parser
  parser.add_argument("src_workbook", help="Enter path and name of src workbook")
  parser.add_argument("tgt_workbook", help="Enter name of the target workbook")
  parser.add_argument("tgt_worksheet", help="Enter name of the target worksheet")
  parser.add_argument("tgt_startcell", help="Enter cell name to start filling in target worksheet")
  parser.add_argument("mapping_file", help="Enter file with sheet mappings")

  # parse the arguments
  args = parser.parse_args()

  createTargetIfNotExists(args.tgt_workbook, args.tgt_worksheet)

  mapping_list = returnMapFileDetails(args.mapping_file)
  #print(mapping_list)

  copyCellsFromSrcToTgt(args.src_workbook, args.tgt_workbook, args.tgt_worksheet, args.tgt_startcell, mapping_list)